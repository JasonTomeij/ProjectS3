import os
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def extract_hu_data():
    # URL van de HU-pagina
    url = "https://www.hu.nl/voltijd-opleidingen/bedrijfskunde/tijdens-de-opleiding"

    # HTML ophalen
    response = requests.get(url)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    # Selecteer alle lijstitems
    items = soup.select(".richtext ol li, .richtext ul li")
    print(f"🔍 Found items: {len(items)}")

    # Alle tekst naar lowercase
    data = [item.get_text(strip=True).lower() for item in items]

    # Lijsten van soft skills en competenties
    SOFT_SKILLS = [
        "communicatie", "schriftelijke communicatie", "mondelinge communicatie",
        "presentatievaardigheden", "onderhandelen", "netwerken", "actief luisteren",
        "klantgerichtheid", "verhalen vertellen", "storytelling", "interpersoonlijke vaardigheden",
        "relatiebeheer", "empathie", "publieke communicatie", "feedback geven", "feedback ontvangen",
        "creativiteit", "out-of-the-box denken", "innovatief denken",
        "visueel denken", "ideeën genereren", "conceptontwikkeling", "branding", "marketingstrategie",
        "copywriting", "contentcreatie", "storytellingvaardigheden", "campagneplanning",
        "analytisch denken", "data-analyse", "probleemoplossend vermogen",
        "datagedreven besluitvorming", "google analytics", "kpi-analyse", "strategisch inzicht",
        "marktanalyse", "onderzoekend vermogen", "meten en evalueren", "resultaatgerichtheid",
        "projectmanagement", "tijdmanagement", "organisatievermogen", "prioriteiten stellen",
        "plannen", "multitasking", "efficiënt werken", "doelgericht werken", "zelfdiscipline",
        "deadline management", "besluitvorming", "strategische planning",
        "samenwerken", "teamwork", "leiderschap", "coaching", "initiatief nemen",
        "betrokkenheid", "conflicthantering", "positieve houding", "zelfreflectie",
        "aanpassingsvermogen", "betrouwbaarheid", "verantwoordelijkheid", "zelfvertrouwen",
        "digitale geletterdheid", "online communicatie", "social media awareness",
        "digitale samenwerking", "digitale marketing", "influencer management",
        "contentstrategie", "data storytelling", "digitale empathie", "ai-vaardigheden",
        "marketingautomatisering", "crm-denken", "growth mindset",
        "ondernemend denken", "commercieel inzicht", "merkdenken",
        "positionering", "consumentenpsychologie", "stakeholdermanagement",
        "budgetbewustzijn", "lange termijn denken", "business development",
        "strategisch communiceren", "onderzoekend vermogen",
        "stressbestendigheid", "doorzettingsvermogen", "flexibiliteit",
        "kritisch denken", "leren leren", "ethisch bewustzijn", "professioneel gedrag",
        "zelfontwikkeling", "open mindedness", "empowerment", "mentale veerkracht",
        "ownership", "klantinzicht", "doelgroepdenken", "klantbeleving",
        "customer journey-denken", "storybranding", "marketingcommunicatie",
        "loyaliteitsdenken", "trendbewustzijn"
    ]

    COMPETENCIES = [
        "strategisch denken", "marktanalyse", "data-analyse", "concurrentieanalyse",
        "probleemanalyse", "onderzoeksvaardigheden", "doelgroepanalyse",
        "besluitvorming", "kritisch denken", "trendonderzoek", "evaluatievaardigheden",
        "kosten-batenanalyse", "risicomanagement", "forecasting", "planningsvaardigheden",
        "branding", "storytelling", "marketingcommunicatie", "public relations",
        "copywriting", "visuele communicatie", "presentatievaardigheden",
        "interne communicatie", "externe communicatie", "multimediale communicatie",
        "contentstrategie", "advertentieplanning", "promotieontwikkeling",
        "digitale marketing", "social media management", "emailmarketing",
        "seo", "sea", "campagnebeheer", "crm-beheer", "webanalyse", "growth hacking",
        "performance marketing", "online adverteren", "digitale strategie",
        "marketingautomatisering", "customer journey mapping", "conversieoptimalisatie",
        "klantgerichtheid", "klantinzicht", "klantrelatiebeheer", "klantbehoud",
        "loyaliteitsmanagement", "customer experience", "doelgroepsegmentatie",
        "service design", "waardepropositieontwikkeling", "marktonderzoek",
        "positionering", "behoefteanalyse", "koopgedraganalyse", "customer lifetime value-denken",
        "projectmanagement", "planning", "organisatievermogen", "tijdmanagement",
        "budgetbeheer", "resourceplanning", "multidisciplinair samenwerken",
        "stakeholdermanagement", "agile werken", "scrum-methodologie",
        "rapportage", "prioriteiten stellen", "kwaliteit bewaken", "operationeel management",
        "creativiteit", "conceptontwikkeling", "ideeëngeneratie", "innovatievermogen",
        "design thinking", "campagneontwikkeling", "probleemoplossend vermogen",
        "visueel denken", "merkstrategie", "prototyping", "trendbewustzijn",
        "empathisch ontwerpen", "user experience", "user interface denken",
        "leiderschap", "teamcoördinatie", "samenwerken", "coaching", "conflicthantering",
        "inspireren", "motiveren", "onderhandelen", "delegeren", "empowerment",
        "initiatief nemen", "zelfreflectie", "besluitvaardigheid", "persoonlijk leiderschap",
        "stressbestendigheid", "aanpassingsvermogen", "doorzettingsvermogen",
        "ethisch handelen", "zelforganisatie", "verantwoordelijkheid nemen",
        "zelfontwikkeling", "leerbereidheid", "resultaatgerichtheid",
        "professioneel gedrag", "integriteit", "ownership", "positieve houding",
        "ondernemerschap", "business development", "financieel inzicht",
        "commercieel inzicht", "ondernemend denken", "budgetbewustzijn",
        "marktgericht handelen", "verkoopvaardigheden", "netwerken",
        "strategisch ondernemerschap", "waardecreatie", "business model innovatie"
    ]

    # 📊 Groeperen per Bron met limiet op lengte
    grouped_results = []
    max_bron_len = 30  # maximale lengte van de bron

    for text in data:
        # Bron inkorten indien nodig
        short_bron = text if len(text) <= max_bron_len else text[:max_bron_len] + "..."

        # soft skills
        for skill in SOFT_SKILLS:
            if skill in text:
                grouped_results.append({"Type": "Soft Skill", "Naam": skill, "Bron": short_bron})
        # competenties
        for comp in COMPETENCIES:
            if comp in text:
                grouped_results.append({"Type": "Competentie", "Naam": comp, "Bron": short_bron})

    # DataFrame aanmaken en duplicaten verwijderen
    df = pd.DataFrame(grouped_results)
    df = df.drop_duplicates(subset=["Type", "Naam", "Bron"]).reset_index(drop=True)

    # Excel pad
    output_path = r"C:\xampp\htdocs\GitHub\Project CMS\Testing with code copies\Real testing\excel_files\Opleiding.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Oude Excel verwijderen
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
        except PermissionError:
            print("⚠️ Sluit eerst het Excel-bestand!")
            return

    # DataFrame naar Excel schrijven
    df.to_excel(output_path, index=False, engine="openpyxl")
    print(f"✅ {len(df)} unieke resultaten opgeslagen in: {output_path}")

    # 🌟 Excel opmaak
    wb = load_workbook(output_path)
    ws = wb.active

    # Headers vet en gecentreerd + kolombreedtes
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        max_length = max(len(str(ws.cell(row=r, column=col).value)) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = max_length + 5

    # Rijen kleuren
    for row in range(2, ws.max_row + 1):
        cell_type = ws.cell(row=row, column=1).value
        fill_color = "ADD8E6" if cell_type == "Soft Skill" else "90EE90"
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    wb.save(output_path)
    print("🎨 Excel-opmaak toegepast!")

if __name__ == "__main__":
    extract_hu_data()
